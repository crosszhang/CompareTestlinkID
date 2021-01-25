""" Scan Agils test spec folder 
Dump the name and testlink id of Spec to result file: Record_SPEC.csv
Andy Zhang 02/23/2020
V1.0
"""


import traceback
import os
import sys
import time
import csv
import openpyxl
import re
import string




csvfile = open('Record_SPEC.csv', 'a', newline='')
csv_write = csv.writer(csvfile, dialect='excel')



def checkfile(filename):
    'check file to find wanted ID'
    if '~' in filename:
        return

    wb = openpyxl.load_workbook(filename)
    ws = wb['Test Overview']
    TestlinkID = ws.cell(row=15,column=2).value
    Author = ws.cell(row=4,column=2).value
    if TestlinkID:
        csv_write.writerow([TestlinkID, filename])



def check_folder(folder):
    'check folder and sub folder to find all wanted files'
    filexternA = "xlsm"
    filexternB = "xlsx"
    list = os.listdir(folder)
    Num = 0

    os.chdir(folder)
#    print('go to next level',folder)
    for i in range(0,len(list)):
        print(list[i])
        if os.path.isfile(list[i]):
            if filexternA in str(list[i]) or filexternB in str(list[i]):
                checkfile(list[i])

        else:
            if 'Obsolete' in str(list[i]):
                print('Skip Obsolete folder')
            else:
                check_folder(list[i])
    os.chdir('..')


def main():


    Record_Time = [time.ctime()]
    csv_write.writerow(Record_Time)
    csv_write.writerow(['TestlinkID','SPEC name'])

    folder = "C:\FW_System_Test\GVL\Charlie_Spec"
#    folder = sys.argv[1]
    check_folder(folder)
    csvfile.close()

if __name__ == "__main__":
    main()








